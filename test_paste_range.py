# -*- coding: utf-8 -*-
"""
Created on Thu Jun 17 09:16:51 2021

@author: Sean Ruane
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side


file_loc = "C:\\Users\\Sean Ruane\\Documents\\"


wb = load_workbook(file_loc + 'phi_summary_raw.xlsx')
ws = wb.active
ws.insert_rows(41, 3)
ws.insert_rows(77, 3)
ws.insert_rows(104, 3)

rows = 3
columns = 14
listab = []

my_blue= openpyxl.styles.colors.Color(rgb='05c3de')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_blue)


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for i in range (1, rows + 1):
    for j in range (1, columns + 1):
        # reading cell value from source excel file
        c = ws.cell(row = i, column = j)
        c.fill = my_fill
        
        # writing the read value to destination excel file
        ws.cell(row = i+40, column = j).value = c.value
        ws.cell(row = i+40, column = j).fill = my_fill
        ws.cell(row = i+40, column = j).border = thin_border
        
        ws.cell(row = i+76, column = j).value = c.value
        ws.cell(row = i+76, column = j).fill = my_fill
        ws.cell(row = i+76, column = j).border = thin_border
        
        ws.cell(row = i+103, column = j).value = c.value
        ws.cell(row = i+103, column = j).fill = my_fill
        ws.cell(row = i+103, column = j).border = thin_border

        
wb.save(file_loc + 'phi_summary_raw617.xlsx')