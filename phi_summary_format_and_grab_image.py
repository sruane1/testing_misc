# -*- coding: utf-8 -*-
"""
Created on Thu Jun 17 11:19:59 2021

@author: Sean Ruane
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
import win32com.client
from PIL import ImageGrab
import os


file_loc = "C:\\Users\\Sean Ruane\\Documents\\"
ipt_name_1 = 'phi_summary_raw.xlsx'
outpt_name = 'phi_summary_raw617.xlsx'

wb = load_workbook(file_loc + ipt_name_1)
ws = wb.active
ws.insert_rows(41, 3)
ws.insert_rows(77, 3)
ws.insert_rows(104, 3)

rows = 3
columns = 14
listab = []

my_blue= openpyxl.styles.colors.Color(rgb='05c3de')
my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_blue)


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for i in range (1, rows + 1):
    for j in range (1, columns + 1):
        # reading cell value from source excel file
        c = ws.cell(row = i, column = j)
        c.fill = my_fill
        
        # writing the read value to destination excel file
        ws.cell(row = i+40, column = j).value = c.value
        ws.cell(row = i+40, column = j).fill = my_fill
        ws.cell(row = i+40, column = j).border = thin_border
        
        ws.cell(row = i+76, column = j).value = c.value
        ws.cell(row = i+76, column = j).fill = my_fill
        ws.cell(row = i+76, column = j).border = thin_border
        
        ws.cell(row = i+103, column = j).value = c.value
        ws.cell(row = i+103, column = j).fill = my_fill
        ws.cell(row = i+103, column = j).border = thin_border

        
wb.save(file_loc + outpt_name)

ipt_name = outpt_name

try:
    wb = load_workbook(file_loc + ipt_name)
    ws = wb.active
    
    o = win32com.client.gencache.EnsureDispatch("Excel.Application")
    wb = o.Workbooks.Open(file_loc + outpt_name) # phi_summary_raw3.xlsx
    ws = wb.Worksheets("sheet1")

    ws.Range(ws.Cells(1, 1), ws.Cells(40, 14)).CopyPicture(Format=win32com.client.constants.xlBitmap)
    img = ImageGrab.grabclipboard()
    imgFile = os.path.join(os.getcwd(), file_loc + 'S1.png')
    img.save(imgFile)

    ws.Range(ws.Cells(41, 1), ws.Cells(76, 14)).CopyPicture(Format=win32com.client.constants.xlBitmap)
    img = ImageGrab.grabclipboard()
    imgFile = os.path.join(os.getcwd(), file_loc + 'S2.png')
    img.save(imgFile)

    ws.Range(ws.Cells(77, 1), ws.Cells(103, 14)).CopyPicture(Format=win32com.client.constants.xlBitmap)
    img = ImageGrab.grabclipboard()
    imgFile = os.path.join(os.getcwd(), file_loc + 'S3.png')
    img.save(imgFile)

    ws.Range(ws.Cells(104, 1), ws.Cells(149, 14)).CopyPicture(Format=win32com.client.constants.xlBitmap)
    img = ImageGrab.grabclipboard()
    imgFile = os.path.join(os.getcwd(), file_loc + 'S4.png')
    img.save(imgFile)

    wb.Close()
except AttributeError:
    # Corner case dependencies.
    import os
    import re
    import sys
    import shutil
    # Remove cache and try again.
    MODULE_LIST = [m.__name__ for m in sys.modules.values()]
    for module in MODULE_LIST:
        if re.match(r'win32com\.gen_py\..+', module):
            del sys.modules[module]
    shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
    from win32com import client
    o = client.gencache.EnsureDispatch('Excel.Application')